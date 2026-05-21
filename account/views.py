from rest_framework import status, generics, viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import PermissionDenied
# Bütün serializer'larımızı tek satırda temizce çağırdık
from .serializers import (
    LoginSerializer, RegisterSerializer, LogoutSerializer, ProfileSerializer,
    UserListSerializer, AdminUserCreateSerializer, AdminUserUpdateSerializer,
)

# Issue 9
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
from .models import MyUser

class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)

        if serializer.is_valid():
            user = serializer.save()
            token, created = Token.objects.get_or_create(user=user)

            return Response(
                {
                    "message": "Kayıt başarılı.",
                    "token": token.key,
                    "user": {
                        "id": user.id,
                        "email": user.email,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "identification_number": user.identification_number,
                        "phone_number": user.phone_number,
                        "address": user.address,
                        "department": user.department.id if user.department else None,
                    },
                },
                status=status.HTTP_201_CREATED,
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)

        if serializer.is_valid():
            user = serializer.validated_data["user"]
            refresh = RefreshToken.for_user(user)

            return Response(
                {
                    "message": "Giriş başarılı.",
                    "tokens": {
                        "access": str(refresh.access_token),
                        "refresh": str(refresh),
                    },
                    "user": {
                        "id": user.id,
                        "email": user.email,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "is_staff": user.is_staff,
                        "is_superuser": user.is_superuser,
                    },
                },
                status=status.HTTP_200_OK,
            )

        return Response(serializer.errors, status=status.HTTP_401_UNAUTHORIZED)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Çıkış başarılı."},
                status=status.HTTP_200_OK
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Sayfanın en üstündeki diğer 'from ...' yazan yerlerin yanına bunu da ekle:

class ProfileAPIView(generics.RetrieveUpdateAPIView):
    serializer_class = ProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user
    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        # Güncellenecek obje (instance) ile isteği atan kişi (request.user) aynı mı kontrolü
        if instance != request.user:
            raise PermissionDenied("Sadece kendi profilinizi güncelleyebilirsiniz.")

        return super().update(request, *args, **kwargs)

# --- YENİ EKLENEN KISIM: Issue #9 ---
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class UserListView(generics.ListAPIView):
    """Deprecated - artik UserViewSet kullaniliyor. Geriye donuk uyumluluk icin tutuluyor."""
    queryset = MyUser.objects.all().order_by('-date_joined')
    serializer_class = UserListSerializer
    permission_classes = [IsAdminUser]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['department', 'is_active', 'is_staff', 'is_superuser']
    search_fields = ['first_name', 'last_name', 'email', 'identification_number']


class UserViewSet(viewsets.ModelViewSet):
    """
    Admin user yonetim ViewSet:
      GET    /api/account/users/         - listele (filter + arama destekli)
      POST   /api/account/users/         - yeni kullanici olustur
      GET    /api/account/users/{id}/    - detay
      PATCH  /api/account/users/{id}/    - guncelle
      DELETE /api/account/users/{id}/    - sil
      POST   /api/account/users/import/  - Excel ile toplu yukle (numara, ad, soyad)
    """
    queryset = MyUser.objects.all().order_by('-date_joined')
    permission_classes = [IsAdminUser]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['department', 'is_active', 'is_staff', 'is_superuser']
    search_fields = ['first_name', 'last_name', 'email', 'identification_number']
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_serializer_class(self):
        if self.action == 'create':
            return AdminUserCreateSerializer
        if self.action in ['update', 'partial_update']:
            return AdminUserUpdateSerializer
        return UserListSerializer

    def get_paginated_or_full(self, queryset):
        page = self.paginate_queryset(queryset)
        if page is not None:
            return self.get_paginated_response(UserListSerializer(page, many=True).data)
        return Response(UserListSerializer(queryset, many=True).data)

    def perform_destroy(self, instance):
        # Admin kendi hesabini silemesin
        if instance.id == self.request.user.id:
            raise PermissionDenied("Kendi hesabinizi silemezsiniz.")
        instance.delete()

    @action(
        detail=False,
        methods=['post'],
        url_path='import',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        """
        Excel dosyasi (.xlsx) ile toplu ogrenci yuklemesi.
        Beklenen kolonlar (header satirini ilk satir olarak okur, basliklarda tirelenir):
          - numara | id_number | identification_number | ogrenci_no | student_no
          - ad     | first_name | isim
          - soyad  | last_name  | soyisim
        Email yoksa <numara>@student.local olarak set edilir, sifre = numara.
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return Response(
                {"detail": "openpyxl yuklu degil. requirements.txt'de bulundugundan emin olun."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        from openpyxl import load_workbook

        file = request.FILES.get('file')
        if not file:
            return Response({"detail": "'file' alaninda bir Excel (.xlsx) dosyasi gonderin."}, status=400)

        # is_staff opsiyonel (egitmen toplu yuklemek icin)
        default_is_staff = str(request.data.get('is_staff', 'false')).lower() in ('1', 'true', 'yes')

        try:
            wb = load_workbook(file, data_only=True, read_only=True)
            ws = wb.active
        except Exception as exc:
            return Response({"detail": f"Excel okunamadi: {exc}"}, status=400)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Dosya bos."}, status=400)

        # Header satirini bul
        headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        col_aliases = {
            "numara": {"numara", "id_number", "identification_number", "ogrenci_no", "ogrenci no", "student_no", "kimlik"},
            "ad": {"ad", "first_name", "isim", "name"},
            "soyad": {"soyad", "last_name", "soyisim", "surname"},
        }

        col_index = {}
        for std_key, aliases in col_aliases.items():
            for idx, h in enumerate(headers):
                if h in aliases:
                    col_index[std_key] = idx
                    break

        # Header bulunamadiysa pozisyonel (numara=0, ad=1, soyad=2) varsay
        if "numara" not in col_index or "ad" not in col_index or "soyad" not in col_index:
            col_index = {"numara": 0, "ad": 1, "soyad": 2}
            data_rows = rows  # tum satirlar veri kabul edilir
        else:
            data_rows = rows[1:]

        created = []
        skipped = []
        errors = []

        for row_idx, row in enumerate(data_rows, start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                numara = str(row[col_index["numara"]] or "").strip()
                ad = str(row[col_index["ad"]] or "").strip()
                soyad = str(row[col_index["soyad"]] or "").strip()

                # Float gibi gelen numarayi ('12345678901.0') temizle
                if numara.endswith('.0'):
                    numara = numara[:-2]

                if not numara or not ad or not soyad:
                    errors.append({"row": row_idx, "error": "Eksik alan (numara/ad/soyad)"})
                    continue
                if len(numara) != 11 or not numara.isdigit():
                    errors.append({"row": row_idx, "numara": numara, "error": "Numara 11 haneli sayi olmali."})
                    continue
                if MyUser.objects.filter(identification_number=numara).exists():
                    skipped.append({"row": row_idx, "numara": numara, "reason": "Zaten mevcut"})
                    continue

                email = f"{numara}@student.local"
                if MyUser.objects.filter(email=email).exists():
                    skipped.append({"row": row_idx, "numara": numara, "reason": f"Email cakisti: {email}"})
                    continue

                user = MyUser(
                    email=email,
                    first_name=ad,
                    last_name=soyad,
                    identification_number=numara,
                    phone_number="",
                    address="",
                    is_staff=default_is_staff,
                    is_superuser=False,
                    is_active=True,
                )
                user.set_password(numara)
                user.save()
                created.append({"row": row_idx, "numara": numara, "id": user.id})
            except Exception as exc:
                errors.append({"row": row_idx, "error": str(exc)})

        return Response({
            "created": len(created),
            "skipped": len(skipped),
            "errors": len(errors),
            "details": {
                "created": created,
                "skipped": skipped,
                "errors": errors,
            },
        }, status=200)
